from playwright.sync_api import sync_playwright
import openpyxl


# xlsx基础操作
class OperateXlsx:
    def __init__(self):
        self.xlsx_address = r'C:\Users\温柔的小茶花\Desktop\样本9\算法挖掘数据样本9-王二-00-00.xlsx'
        self.sheet_name = '王二'
        self.poi_id_column = 'a'
        self.poi_name_column = 'b'
        self.reference_point_column = 'e'
        self.reference_fence_column = 'f'
        self.correct_fence_column = 'j'
        self.wb = openpyxl.load_workbook(self.xlsx_address)
        self.ws = self.wb[self.sheet_name]

    def visit_poi_id(self, row_num):
        return self.ws[f'{self.poi_id_column}{row_num}'].value

    def visit_poi_name(self, row_num):
        return self.ws[f'{self.poi_name_column}{row_num}'].value

# 访问参考围栏, 如参考围栏为None, 访问参考位置
    def visit_reference_point_or_fence(self, row_num):
        fence_value = self.ws[f'{self.reference_fence_column}{row_num}'].value
        if fence_value == r'\N':
            return fence_value
        else:
            return self.ws[f'{self.reference_point_column}{row_num}'].value

# 写入'正确围栏'数据
    def write_correct_reference(self, row_num, correct_reference_data):
        self.ws[f'{self.correct_fence_column}{row_num}'] = correct_reference_data
        self.wb.save(self.xlsx_address)


# 画图工具
def draw_pic(reference_fence):
    # 1/ tool地址栏清空内容
    input_ele = page.locator('xpath=//*[@id="wkt"]')
    input_ele.fill('')
    # 2/ 输入'参考围栏'
    input_ele.fill(reference_fence)
    # 3/ 点击' draw it'
    draw_it_ele = page.locator('xpath=//*[@id="draw"]')
    draw_it_ele.click()
    # 4/ 点击'鼠标绘制面'
    mouse_draw_plane_ele = page.locator('xpath=// *[ @ id = "polygon"]')
    mouse_draw_plane_ele.click()
    # 5/ 等待用户绘制图像
    page.locator('xpath=//*[@id="iCenter"]/div[1]/div/div[1]/div/div/div/div').wait_for(timeout=1000000, state='visible')
    # 6/ 获取 用户绘制图像产生的数据
    input_content = page.eval_on_selector('xpath=//*[@id="wkt"]', "(element) => element.value")
    return input_content


a = OperateXlsx()
row_num_1 = int(input('起始数:\n'))

with sync_playwright() as p:
    # 创建一个浏览器实例
    browser = p.chromium.launch(headless=False)
    # 创建含登录状态的浏览器上下文
    context = browser.new_context(viewport={'width': 1440, 'height': 900})

    # 创建一个page对象
    page = context.new_page()
    while True:
        try:
            # 转到页面
            page.goto('https://page.cainiao.com/cntms/cnlbs/cnlbs_visual/cnlbs_visual.html')
            # 打印当前行号及地址
            show_poi_name = a.visit_poi_name(row_num_1)
            print(f'{row_num_1}, {show_poi_name}')

            # 1/ 从表格中获取参考围栏信息
            input_information = a.visit_reference_point_or_fence(row_num_1)
            # 2/ 在画图工具中产生输出信息
            output_information = draw_pic(input_information)
            # 3/ 把输出信息写入到表格中
            a.write_correct_reference(row_num_1, output_information)
        except:
            # (如果画不出图像就直接x掉浏览器--关闭page对象,产生except，然后)创建一个新的page对象
            page = context.new_page()

        row_num_1 += 1
