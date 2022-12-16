"""
通过行政区域代码获取地区信息.   行政区域代码excel已经保存在目录中.
"""
import openpyxl


wb = openpyxl.load_workbook('行政区划代码.xlsx')
ws = wb['Sheet']
dict1 = {}
for i in range(4, 3215):
    dict1[ws['a' + str(i)].value] = ws['b' + str(i)].value

wb1 = openpyxl.load_workbook('算法挖掘数据样本9-王二-00-00.xlsx')
ws1 = wb1['王二']
for i in range(1, 5002):
    city_code = str(ws1['d' + str(i)].value)
    poi_name = ws1['b' + str(i)].value
    try:
        city_name = dict1[city_code]
        ws1['l' + str(i)] = city_name + poi_name
        print(city_code, city_name)
    except:
        ws1['l' + str(i)] = poi_name

wb1.save('算法挖掘数据样本9-王二-00-00.xlsx')
