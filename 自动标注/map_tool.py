import logging
import webbrowser
import openpyxl
import requests
import re
from time import strftime

logging.basicConfig(level=logging.INFO, format='\n%(asctime)s-%(levelname)s: %(message)s\n')


class 表格():
    def __init__(self, 工作簿地址=None, 工作表名称=None, poiid所在列=None):
        self.工作簿地址 = 工作簿地址
        self.工作表名称 = 工作表名称
        if self.工作簿地址 == None:
            self.工作簿地址 = input('\n请输入工作簿地址:\t')
            self.工作表名称 = input('\n请输入工作表名称:\t')
        self.workbook_name = str(self.工作簿地址)[0:-6]
        self.工作簿 = openpyxl.load_workbook(self.工作簿地址)
        self.工作表 = self.工作簿[self.工作表名称]
        self.poiid所在列 = poiid所在列
        if self.poiid所在列 == None:
            self.poiid所在列 = input('\n请输入poiid所在列:\t')

    def 选定单元格(self, 列数=None, 行数=None):
        try:
            if 列数 == None:
                列数 = input('\n请输入单元格的列数(字母形式):\t')
            if 行数 == None:
                行数 = input('\n请输入单元格的行数:\t')
            return self.工作表[f'{列数}{行数}']
        except:
            logging.error(f'选定单元格{列数}{行数}出错!')

    def 打印单元格(self, 列数=None, 行数=None):
        单元格 = 表格.选定单元格(self, 列数=列数, 行数=行数)
        logging.info(f'正在打印单元格 {列数.upper()}{行数}')
        try:
            print(单元格.value)
        except:
            logging.error('打印单元格出错!')

    def 写入单元格(self, 列数=None, 行数=None, 数据=None):
        if 数据 == None:
            数据 = input('\n请输入需要写入的数据:\t')
        logging.info(f'正在将"{数据}"写入单元格 {列数.upper()}{行数}')
        try:
            单元格 = 表格.选定单元格(self, 列数=列数, 行数=行数)
            单元格.value = 数据
            表格.save_workbook()
        except:
            logging.error('数据写入单元格出错!')

    def 写入单元格_不打印信息(self, 列数=None, 行数=None, 数据=None):
        if 数据 == None:
            数据 = input('\n请输入需要写入的数据:\t')
        try:
            单元格 = 表格.选定单元格(self, 列数=列数, 行数=行数)
            单元格.value = 数据
            self.工作簿.save(self.工作簿地址)
        except:
            logging.error('数据写入单元格出错!')

    def save_workbook(self):
        time_format = strftime('%d__%H_%M')
        backups_workbook_name = f'{time_format}__{self.workbook_name}.xlsx'
        self.工作簿.save(backups_workbook_name)
        logging.info('成功备份文件!')

    def save_workbook_2(self):
        list_name = re.split('-', self.workbook_name)
        list_name[2] = str(int(list_name[2])+400)
        list_name[3] = str(int(list_name[3])+400)
        self.workbook_name = f'{list_name[0]}-{list_name[1]}-{list_name[2]}-{list_name[3]}'
        backups_workbook_name = f'{self.workbook_name}.xlsx'
        self.工作簿.save(backups_workbook_name)
        logging.info('成功备份文件!')


class 表格_配置用(表格):
    def __init__(self, 工作簿地址=None, 工作表名称=None):
        self.工作簿地址 = 工作簿地址
        self.工作表名称 = 工作表名称
        if self.工作簿地址 == None:
            self.工作簿地址 = input('\n请输入工作簿地址:\t')
            self.工作表名称 = input('\n请输入工作表名称:\t')
        self.工作簿 = openpyxl.load_workbook(self.工作簿地址)
        self.工作表 = self.工作簿[self.工作表名称]
        self.workbook_name = str(self.工作簿地址)[0:-5]


class 高德工具(表格):
    def __init__(self, 工作簿地址=None, 工作表名称=None, poiid所在列=None, ):
        self.工作簿地址 = 工作簿地址
        self.工作表名称 = 工作表名称
        if self.工作簿地址 == None:
            self.工作簿地址 = input('\n请输入工作簿地址:\t')
            self.工作表名称 = input('\n请输入工作表名称:\t')
        self.工作簿 = openpyxl.load_workbook(self.工作簿地址)
        self.工作表 = self.工作簿[self.工作表名称]
        self.poiid所在列 = poiid所在列
        if self.poiid所在列 == None:
            poiid所在列 = input('\n请输入poiid所在列:\t')

    def 获取带市区的aoi名字(self, poi_id):
        URL = 'https://restapi.amap.com/v5/place/detail?parameters'
        params = {'key': '66c3cd09c85e728fc3f769cbd05e63c0', 'id': poi_id}
        HTML = requests.get(url=URL, params=params)
        数据 = HTML.json()
        if 数据['pois'] != []:
            名字 = 数据['pois'][0]['name']
            城市 = 数据['pois'][0]['cityname']
            县区 = 数据['pois'][0]['adname']
            经纬度 = 数据['pois'][0]['location']
            带地区的aoi名字 = 城市 + 县区 + 名字
            带地区的aoi名字 = 带地区的aoi名字.strip()
        else:
            带地区的aoi名字 = ''
        return 带地区的aoi名字

    def 调起高德地图(self, poiid所在列=None, 行数=None):
        if poiid所在列 == None:
            poiid所在列 = self.poiid所在列
            poiid所在列 = poiid所在列.upper()
        if 行数 == None:
            行数 = input('\n请输入需要调起高德地图的行数:\t')
            if 行数 == 'b':
                return 行数
        url = 高德工具.选定单元格(self, 列数=poiid所在列, 行数=行数)  # 这里url只是 poiid的单元格
        url = f'https://ditu.amap.com/place/{url.value}'
        webbrowser.open(url)
        print(f'已调起"{行数}"行的高德地图.')

    def 批量调起高德地图(self, 起始数=None, 行数=None):
        while True:
            if 起始数 == None:
                起始数 = input(f'\n请输入批量调起地图的起始数:\t')
                起始数 = int(起始数)
            i = 起始数 - 10
            while True:
                i += 10
                i2 = i + 10
                for i3 in range(i, i2):
                    高德工具.调起高德地图(self, poiid所在列=self.poiid所在列, 行数=i3)
                logging.info(f'已打开{i}到{i2 - 1}行的高德地图')
                选择操作 = input('回车以继续,输入数字重新选择操作:\t')
                if 选择操作 != '':
                    return 选择操作


class 百度工具(表格):
    def __init__(self, 工作簿地址=None, 工作表名称=None, poiid所在列=None):
        self.工作簿地址 = 工作簿地址
        self.工作表名称 = 工作表名称
        self.poiid所在列 = poiid所在列
        if self.工作簿地址 == None:
            self.工作簿地址 = input('\n请输入工作簿地址:\t')
            self.工作表名称 = input('\n请输入工作表名称:\t')
        self.工作簿 = openpyxl.load_workbook(self.工作簿地址)
        self.工作表 = self.工作簿[self.工作表名称]

    def 调起百度地图(self, 行数=None):
        if 行数 == None:
            行数 = input('\n请输入需要调起百度地图的行数:\t')
            if 行数 == 'b':
                return 行数
        高德工具_暂用 = 高德工具(工作簿地址=self.工作簿地址, 工作表名称=self.工作表名称, poiid所在列=self.poiid所在列)
        poi_id = 高德工具_暂用.选定单元格(行数=行数, 列数=self.poiid所在列).value
        地址 = 高德工具_暂用.获取带市区的aoi名字(poi_id=poi_id)
        if 地址 == '' or 地址 == None:
            地址 = '无'
            print('当前地址在百度地图上搜索不到')
        if 地址 != '无':
            百度地图链接 = f'http://api.map.baidu.com/geocoder?address={地址}&output=html&src=webapp.baidu.openAPIdemo'
            webbrowser.open(百度地图链接)
            print(f'已调起"{地址}"的百度地图.')

    def 批量调起百度地图(self, 起始数=None, ):
        while True:
            if 起始数 == None:
                起始数 = input(f'\n请输入批量调起地图的起始数:\t')
                起始数 = int(起始数)
            i = 起始数 - 10
            while True:
                i += 10
                i2 = i + 10
                for i3 in range(i, i2):
                    百度工具.调起百度地图(self, 行数=i3)
                logging.info(f'已打开{i}到{i2 - 1}行的百度地图')
                选择操作 = input('回车以继续,输入数字重新选择操作')
                if 选择操作 != '':
                    return 选择操作


# ~ 1\从"配置.xlsx"获取信息
配置表 = 表格_配置用(工作簿地址="配置.xlsx", 工作表名称="Sheet1")
工作簿名称 = 配置表.选定单元格('b', 1).value
工作表名称 = 配置表.选定单元格('b', 2).value


暂用表 = 表格_配置用(工作簿地址=工作簿名称, 工作表名称=工作表名称)
for 列名 in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
    单元格 = 暂用表.选定单元格(f'{列名}', 1).value
    if 单元格 == 'poiid':
        poiid所在列 = 列名
    if 单元格 == '是否正确':
        是否正确所在列 = 列名
    if 单元格 == '正确围栏':
        正确围栏所在列 = 列名
    if 单元格 == '错误原因':
        错误原因所在列 = 列名

# ~ 2\信息传给两个工具
tool = 高德工具(工作簿地址=工作簿名称, 工作表名称=工作表名称, poiid所在列=poiid所在列)
tool1 = 百度工具(工作簿地址=工作簿名称, 工作表名称=工作表名称, poiid所在列=poiid所在列)
# ~ 3\选择操作
logging.info(
    f'欢迎使用王二的打标辅助工具v1.1~\n\n当前操作的工作簿: "{工作簿名称}"\t当前操作的工作表: "{工作表名称}"')
暂用表.save_workbook()
print('0\备份工作簿(请先保存当前工作簿)')
print('1\调起10个高德地图')
print('2\调起10个百度地图')
print('3\单个调起高德地图')
print('4\单个调起百度地图')
print('b\在任意输入位置返回菜单')
print('q\退出程序')

选择操作 = input('\n请选择操作:\t')
# ~ 4\执行操作		退出按钮			返回选择按钮
while True:
    if 选择操作 == '0':
        暂用表 = 表格_配置用(工作簿地址=工作簿名称, 工作表名称=工作表名称)
        暂用表.save_workbook_2()
        选择操作 = input('\n输入数字重新选择操作:\t')
        continue
        起始数 = int(起始数)
        选择操作 = tool.批量调起高德地图(起始数=起始数)
        continue
    if 选择操作 == '1':
        起始数 = input('\n请输入批量调起的起始数:\t')
        if 起始数 == 'b':
            选择操作 = input('\n输入数字重新选择操作:\t')
            continue
        起始数 = int(起始数)
        选择操作 = tool.批量调起高德地图(起始数=起始数)
        continue
    if 选择操作 == '2':
        起始数 = input('\n请输入批量调起的起始数:\t')
        if 起始数 == 'b':
            选择操作 = input('\n输入数字重新选择操作:\t')
            continue
        起始数 = int(起始数)
        选择操作 = tool1.批量调起百度地图(起始数=起始数)
        continue
    if 选择操作 == '3':
        是否返回上级菜单 = tool.调起高德地图()
        if 是否返回上级菜单 == 'b':
            选择操作 = input('\n输入数字重新选择操作:\t')
            continue
    if 选择操作 == '4':
        是否返回上级菜单 = tool1.调起百度地图()
        if 是否返回上级菜单 == 'b':
            选择操作 = input('\n输入数字重新选择操作:\t')
            continue
    if 选择操作 == 'b':
        选择操作 = input('\n输入数字重新选择操作:\t')
        continue
    if 选择操作 == 'q':
        exit(0)
    if 选择操作 not in ['1', '2', '3', '4', 'b', 'q']:
        logging.error('输入有误')
        选择操作 = input('\n输入数字重新选择操作:\t')
        continue
