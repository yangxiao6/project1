import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait

'''
1/ 简单的api
2/ 复杂的api----由简单的api组合而成
3/ "poi_id" & "poi_name" & "参考围栏" & "正确围栏" 所在列都在程序开头手动获取
4/ 始终只用一个xlsx
5/ 工作簿名称更新 由另一个exe执行
'''


# xlsx基础操作 用一个类, 不同的表格操作设计为不同的类方法
class OperateXlsx:
    def __init__(self):

        self.xlsx_address = '算法挖掘数据样本7-王二-00-00.xlsx'
        self.sheet_name = '王二'
        self.poi_id_column = 'a'
        self.poi_name_column = 'b'
        self.reference_point_column = 'e'
        self.reference_fence_column = 'f'
        self.correct_fence_column = 'j'
        self.wb = openpyxl.load_workbook(self.xlsx_address)
        self.ws = self.wb[self.sheet_name]

    def visit_poi_id(self, row_num):
        return self.ws[f'{self.poi_id_column}{row_num}'].value

    def visit_poi_name(self, row_num):
        return self.ws[f'{self.poi_name_column}{row_num}'].value

    def visit_reference_point_or_fence(self, row_num):
        fence_value = self.ws[f'{self.reference_fence_column}{row_num}'].value
        if fence_value is not None:
            return fence_value
        else:
            return self.ws[f'{self.reference_point_column}{row_num}'].value

    def write_correct_reference(self, row_num, correct_reference_data):
        self.ws[f'{self.correct_fence_column}{row_num}'] = correct_reference_data
        self.wb.save(self.xlsx_address)



# 画图工具 用一个类, 不同的
"""
1/ 打开浏览器
2/ 转到画图工具的url
3/ 输入'参考围栏'
4/ 点击' draw it'
5/ 点击'鼠标绘制面'
6/ 等待用户绘制图像
7/ 获取 用户绘制图像产生的数据  /   获取 用户放弃绘制图像的信号
8/ 将数据填入excel文件对应的单元格(如7中是放弃绘制,则数据为空)
9/ 返回到3 输入'参考围栏'
"""


class DrawPic:
    def __init__(self):
        self.browser = webdriver.Chrome()
        self.browser.maximize_window()  # 最大化浏览器窗口
        self.browser.get('https://page.cainiao.com/cntms/cnlbs/cnlbs_visual/cnlbs_visual.html')  # 1/打开浏览器 2/转到画图工具的url
        self.input_ele = self.browser.find_element(By.XPATH, '//*[@id="wkt"]')
        self.draw_it_ele = self.browser.find_element(By.XPATH, '//*[@id="draw"]')
        self.mouse_draw_plane_ele = self.browser.find_element(By.XPATH, '// *[ @ id = "polygon"]')

    def draw_pic(self, reference_fence):
        self.input_ele.clear()  # 清除输入框中原有内容
        self.input_ele.send_keys(reference_fence)  # 3/ 输入'参考围栏'
        self.draw_it_ele.click()  # 4/ 点击' draw it'
        self.mouse_draw_plane_ele.click()  # 5/ 点击'鼠标绘制面'
        WebDriverWait(driver=self.browser, timeout=6000).until(  # 6/ 等待用户绘制图像
            ec.presence_of_element_located(('xpath', '//*[@id="iCenter"]/div[1]/div/div[1]/div/div/div/div')))
        input_content = self.input_ele.get_attribute('value')  # 7/ 获取 用户绘制图像产生的数据
        return input_content
# 7/ 获取 用户绘制图像产生的数据  /   获取 用户放弃绘制图像的信号
# 8/ 将数据填入excel文件对应的单元格(如7中是放弃绘制,则数据为空)
# 9/ 返回到3 输入'参考围栏'


a = OperateXlsx()
b = DrawPic()
row_num_1 = int(input('起始数:\n'))

while True:
    try:
        input_information = a.visit_reference_point_or_fence(row_num_1)             # 1/ 从表格中获取参考围栏信息
        show_poi_name = a.visit_poi_name(row_num_1)                                 # 1.1/ 显示poi_name
        print(f'{row_num_1}, {show_poi_name}')
        output_information = b.draw_pic(input_information)                          # 2/ 在画图工具中产生输出信息
        a.write_correct_reference(row_num_1, output_information)                    # 3/ 把输出信息写入到表格中
        row_num_1 += 1
    except:
        b = DrawPic()
        row_num_1 += 1
