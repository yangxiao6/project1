import openpyxl
from playwright.sync_api import sync_playwright
import time
import json
import re

'''
1/ 简单的api
2/ 复杂的api----由简单的api组合而成
3/ "poi_id" & "poi_name" & "参考围栏" & "正确围栏" 所在列都在程序开头手动获取
4/ 始终只用一个xlsx
5/ 工作簿名称更新 由另一个exe执行
'''


# xlsx基础操作 用一个类, 不同的表格操作设计为不同的类方法
class OperateXlsx:
    def __init__(self):

        self.xlsx_address = '算法挖掘数据样本8-王二-00-00.xlsx'
        self.sheet_name = '王二'
        self.poi_id_column = 'a'
        self.poi_name_column = 'b'
        self.reference_point_column = 'e'
        self.reference_fence_column = 'f'
        self.correct_fence_column = 'j'
        self.wb = openpyxl.load_workbook(self.xlsx_address)
        self.ws = self.wb[self.sheet_name]

    def visit_poi_id(self, row_num):
        return self.ws[f'{self.poi_id_column}{row_num}'].value

    def visit_poi_name(self, row_num):
        return self.ws[f'{self.poi_name_column}{row_num}'].value

    def visit_reference_point_or_fence(self, row_num):
        fence_value = self.ws[f'{self.reference_fence_column}{row_num}'].value
        if fence_value is not None:
            return fence_value
        else:
            return self.ws[f'{self.reference_point_column}{row_num}'].value

    def write_correct_reference(self, row_num, correct_reference_data):
        self.ws[f'{self.correct_fence_column}{row_num}'] = correct_reference_data
        self.wb.save(self.xlsx_address)


# 画图工具 用一个类, 不同的
def draw_pic(reference_fence):
    # 2/ tool地址栏清空内容
    input_ele.fill('')
    # 3/ 输入'参考围栏'
    input_ele.fill(reference_fence)
    # 4/ 点击' draw it'
    draw_it_ele.click()
    # 5/ 点击'鼠标绘制面'
    mouse_draw_plane_ele.click()
    # 6/ 等待用户绘制图像
    page.locator('xpath=//*[@id="iCenter"]/div[1]/div/div[1]/div/div/div/div').wait_for(timeout=1000000, state='visible')
    # 7/ 获取 用户绘制图像产生的数据
    input_content = page.eval_on_selector('xpath=//*[@id="wkt"]', "(element) => element.value")
    return input_content


# 导入openpyxl库
import openpyxl
# 打开文件
import requests

wb = openpyxl.load_workbook('算法挖掘数据样本8-王二-00-00.xlsx')               # workbook,工作簿
ws = wb['王二']                                                   # worksheet,工作表
poiid_column = 'a'                                               # poiid所在列
reference_point_column = 'e'                                     # 参考位置所在列
reference_fence_column = 'f'                                     # 参考围栏所在列
query_row = '21'                                                 # 查询的行数


# 通过高德api获取完整的aoi名字
def 获取带市区的aoi名字(poi_id):
    URL = 'https://restapi.amap.com/v5/place/detail?parameters'
    params = {'key': '66c3cd09c85e728fc3f769cbd05e63c0', 'id': poi_id}
    HTML = requests.get(url=URL, params=params)
    数据 = HTML.json()
    if 数据['pois']:
        名字 = 数据['pois'][0]['name']
        城市 = 数据['pois'][0]['cityname']
        县区 = 数据['pois'][0]['adname']
        经纬度 = 数据['pois'][0]['location']
        带地区的aoi名字 = 城市 + 县区 + 名字
        带地区的aoi名字 = 带地区的aoi名字.strip()
    else:
        带地区的aoi名字 = ''
    return 带地区的aoi名字

# 通过高德地图url获取完整的aoi名字
# def 从地图页获取aoi名字(poi_id):


a = OperateXlsx()
row_num_1 = int(input('起始数:\n'))

with sync_playwright() as p:
    # 创建一个浏览器实例
    browser = p.chromium.launch(headless=False)
    # 创建含登录状态的浏览器上下文
    with open("baidu_state.json") as f:
        storage_state = json.loads(f.read())
    context1 = browser.new_context(storage_state=storage_state)
    context = browser.new_context(viewport={'width': 1536, 'height': 864})

    # 创建一个page对象
    page1 = context1.new_page()
    page = context.new_page()
    while True:
        try:
            # 获取poiid
            poiid = a.visit_poi_id(row_num_1)
            amap_url = f'https://ditu.amap.com/place/{poiid}'
            address = 获取带市区的aoi名字(poiid)
            baidu_url = f'http://api.map.baidu.com/geocoder?address={address}&output=html&src=webapp.baidu.openAPIdemo'
            # 转到页面
            page1.goto(baidu_url)
            page.goto('https://page.cainiao.com/cntms/cnlbs/cnlbs_visual/cnlbs_visual.html')

            input_ele = page.locator('xpath=//*[@id="wkt"]')
            draw_it_ele = page.locator('xpath=//*[@id="draw"]')
            mouse_draw_plane_ele = page.locator('xpath=// *[ @ id = "polygon"]')

            input_information = a.visit_reference_point_or_fence(row_num_1)  # 1/ 从表格中获取参考围栏信息
            show_poi_name = a.visit_poi_name(row_num_1)  # 1.1/ 显示poi_name
            print(f'{row_num_1}, {show_poi_name}')
            # 2/ 在画图工具中产生输出信息----------------------------------
            output_information = draw_pic(input_information)
            a.write_correct_reference(row_num_1, output_information)  # 3/ 把输出信息写入到表格中
            row_num_1 += 1
        except:
            row_num_1 += 1
            # 创建一个page对象
            page = context.new_page()
